#logistic_map_demo.py

import os
import sys
import tempfile
import openpyxl
import numpy as np
import pyqtgraph as pg
import polars as pl
import pandas as pd

from openpyxl.drawing.image import Image
from pyqtgraph.exporters import ImageExporter

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout,
    QHBoxLayout, QSlider, QLabel, QPushButton, QFileDialog,
    QMessageBox, QLineEdit
)
from PyQt6.QtCore import Qt, QTimer
from PyQt6.QtGui import QPixmap, QGuiApplication, QAction, QDoubleValidator


class LogisticMapDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Демонстрация теории хаоса: логистическое отображение")
        self.setGeometry(100, 100, 1400, 700)

        self.animating = False
        self.animation_direction = 1  # 1 - увеличение r, -1 - уменьшение
        self.timer = QTimer()
        self.timer.timeout.connect(self.animation_step)
        self.current_x0 = 0.5  # начальное условие
        
        # Данные для хранения истории
        self.history_r = []
        self.history_x = []
        self.current_r = 3.2
        self.all_iterations = []  # для экспорта всех данных
        
        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QHBoxLayout(central_widget)

        # Левая панель с графиками
        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)

        # График 1: итерационная диаграмма
        self.plot_iter = pg.PlotWidget(title="Итерации (xn от n)")
        self.plot_iter.setLabel('left', 'xn')
        self.plot_iter.setLabel('bottom', 'Номер итерации')
        self.plot_iter.setXRange(0, 200)
        self.plot_iter.setYRange(0, 1)
        left_layout.addWidget(self.plot_iter)

        # График 2: бифуркационная диаграмма
        self.plot_bifur = pg.PlotWidget(title="Бифуркационная диаграмма")
        self.plot_bifur.setLabel('left', 'x')
        self.plot_bifur.setLabel('bottom', 'r')
        self.plot_bifur.setXRange(2.5, 4.0)
        self.plot_bifur.setYRange(0, 1)
        left_layout.addWidget(self.plot_bifur)

        main_layout.addWidget(left_panel, stretch=3)

        # Правая панель с управлением
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        
        # --- Блок ввода начального условия ---
        self.label_x0 = QLabel("Начальное значение x₀:")
        self.label_x0.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(self.label_x0)
        
        self.x0_input = QLineEdit("0.5")
        self.x0_input.setValidator(QDoubleValidator(0.01, 0.99, 3))  # от 0.01 до 0.99
        self.x0_input.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.x0_input.textChanged.connect(self.on_x0_changed)
        right_layout.addWidget(self.x0_input)

        # --- Блок управления анимацией ---
        self.btn_animate = QPushButton("▶ Запустить анимацию")
        self.btn_animate.clicked.connect(self.toggle_animation)
        right_layout.addWidget(self.btn_animate)

        self.animation_speed = QSlider(Qt.Orientation.Horizontal)
        self.animation_speed.setMinimum(1)
        self.animation_speed.setMaximum(20)
        self.animation_speed.setValue(10)
        self.animation_speed.setTickInterval(5)
        self.animation_speed.setToolTip("Скорость анимации")
        right_layout.addWidget(QLabel("Скорость:"))
        right_layout.addWidget(self.animation_speed)

        self.label_r = QLabel("r = 3.200")
        self.label_r.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(self.label_r)

        self.slider = QSlider(Qt.Orientation.Vertical)
        self.slider.setMinimum(250)
        self.slider.setMaximum(400)
        self.slider.setValue(320)  # 3.20
        self.slider.setTickInterval(10)
        self.slider.setTickPosition(QSlider.TickPosition.TicksRight)
        self.slider.valueChanged.connect(self.on_slider_change)
        right_layout.addWidget(self.slider)

        self.label_lyap = QLabel("Показатель Ляпунова: вычисляется...")
        self.label_lyap.setAlignment(Qt.AlignmentFlag.AlignCenter)
        right_layout.addWidget(self.label_lyap)
        
        # Кнопки экспорта
        self.btn_export_png = QPushButton("📸 Экспорт графиков в PNG")
        self.btn_export_png.clicked.connect(self.export_graphs_png)
        right_layout.addWidget(self.btn_export_png)
        
        self.btn_export_excel = QPushButton("📊 Экспорт данных в Excel")
        self.btn_export_excel.clicked.connect(self.export_data_excel)
        right_layout.addWidget(self.btn_export_excel)
        
        self.btn_export_all = QPushButton("📈 Экспорт графиков в Excel")
        self.btn_export_all.clicked.connect(self.export_graphs_to_excel)
        right_layout.addWidget(self.btn_export_all)

        right_layout.addStretch()
        main_layout.addWidget(right_panel, stretch=1)

        # Данные для бифуркационной диаграммы
        self.bifur_r = []   # значения r
        self.bifur_x = []   # соответствующие x после переходного процесса
        self.bifur_scatter = pg.ScatterPlotItem(size=2, brush='w')
        self.plot_bifur.addItem(self.bifur_scatter)
        
        # Меню
        self.create_menu()

        # Начальное построение
        self.update_plots()
        
        self.animation_speed.valueChanged.connect(self.update_animation_speed)
        
    def update_animation_speed(self, value):
        if self.animating:
            self.timer.setInterval(int(100 / value))

    def on_x0_changed(self, text):
        """Обработчик изменения x₀"""
        try:
            x0 = float(text)
            if 0 < x0 < 1:
                self.current_x0 = x0
                self.update_plots()  # пересчитать с новым x0
            else:
                self.statusBar().showMessage("x₀ должно быть между 0 и 1")
        except ValueError:
            pass  # игнорируем нечисловой ввод
        
    def create_menu(self):
        menubar = self.menuBar()
        file_menu = menubar.addMenu("Файл")
        
        export_png_action = QAction("Экспорт графиков в PNG", self)
        export_png_action.triggered.connect(self.export_graphs_png)
        file_menu.addAction(export_png_action)
        
        export_excel_action = QAction("Экспорт данных в Excel", self)
        export_excel_action.triggered.connect(self.export_data_excel)
        file_menu.addAction(export_excel_action)        
    
    def on_slider_change(self):
        self.update_plots()

    def update_plots(self):
        # Получаем текущее r (слайдер выдаёт значения от 250 до 400 -> делим на 100)
        r = self.slider.value() / 100.0
        self.current_r = r
        self.label_r.setText(f"r = {r:.3f}")

        # --- Итерационная диаграмма ---
        n_iter = 500   # общее число итераций
        x0 = self.current_x0       # начальное условие
        x = np.zeros(n_iter)
        x[0] = x0
        for i in range(1, n_iter):
            x[i] = r * x[i-1] * (1 - x[i-1])
            
        # Сохраняем для экспорта
        self.all_iterations = x.tolist()

        # Отображаем последние 200 итераций
        self.plot_iter.clear()
        self.plot_iter.plot(x[-200:], pen='y')

        # --- Обновление бифуркационной диаграммы ---
        # Для текущего r сохраняем значения x после переходного процесса (последние 100 итераций)
        transient = 400
        stable_x = x[transient:]   # значения после переходного процесса

        # Добавляем точки для этого r
        # Чтобы не перегружать график, добавляем не все точки, а каждую 2-ю (опционально)
        for val in stable_x[::2]:   # берём каждую вторую для наглядности
            self.bifur_r.append(r)
            self.bifur_x.append(val)
            self.history_r.append(r)
            self.history_x.append(val)

        # Обновляем scatter-график
        self.bifur_scatter.setData(self.bifur_r, self.bifur_x)

        # --- Вычисление показателя Ляпунова (упрощённо) ---
        lyap = self.compute_lyapunov(r, x0)
        self.label_lyap.setText(f"Показатель Ляпунова: {lyap:.4f}")

    def compute_lyapunov(self, r, x0, n=500):
        """
        Приближённое вычисление показателя Ляпунова для логистического отображения.
        λ = lim (1/n) Σ ln |f'(x_i)|
        f'(x) = r * (1 - 2x)
        """
        x = x0
        sum_ln = 0.0
        for i in range(n):
            x = r * x * (1 - x)
            # Производная
            df = abs(r * (1 - 2 * x))
            if df > 0:
                sum_ln += np.log(df)
        return sum_ln / n
    
    def export_graphs_png(self):
        """Экспорт графиков в PNG"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить графики", "", "PNG Files (*.png)"
        )
        if file_path:
            try:
                # Экспорт первого графика
                exporter1 = ImageExporter(self.plot_iter.plotItem)
                exporter1.parameters()['width'] = 800
                base_path = file_path.rsplit('.', 1)[0]
                exporter1.export(f'{base_path}_iter.png')
                
                # Экспорт второго графика
                exporter2 = ImageExporter(self.plot_bifur.plotItem)
                exporter2.parameters()['width'] = 800
                exporter2.export(f'{base_path}_bifur.png')
                
                self.statusBar().showMessage(f"Графики сохранены: {base_path}_iter.png и {base_path}_bifur.png")
            except Exception as e:
                self.statusBar().showMessage(f"Ошибка экспорта: {str(e)}")
            
    def export_data_excel(self):
        """Экспорт данных в Excel с использованием pandas"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить данные", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            # Исправлено: pd.DataFrame (было pdDataFrame)
            df_iterations = pd.DataFrame({
                "iteration": list(range(len(self.all_iterations))),
                "x_value": self.all_iterations,
                "r_parameter": [self.current_r] * len(self.all_iterations)
            })
            
            # Создаем данные для бифуркационной диаграммы
            df_bifur = pd.DataFrame({
                "r_parameter": self.history_r[-1000:] if len(self.history_r) > 1000 else self.history_r,
                "x_value": self.history_x[-1000:] if len(self.history_x) > 1000 else self.history_x
            })
            
            # Экспортируем в Excel
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df_iterations.to_excel(writer, sheet_name='Итерации', index=False)
                df_bifur.to_excel(writer, sheet_name='Бифуркация', index=False)
            
            self.statusBar().showMessage(f"Данные сохранены: {file_path}") 
            
    def export_graphs_to_excel(self):
        """Экспорт графиков в Excel как встроенные изображения"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить графики в Excel", "", "Excel Files (*.xlsx)"
        )
        if file_path:
            # Создаем временные файлы для графиков
            with tempfile.NamedTemporaryFile(suffix='_iter.png', delete=False) as tmp_iter:
                iter_path = tmp_iter.name
            with tempfile.NamedTemporaryFile(suffix='_bifur.png', delete=False) as tmp_bifur:
                bifur_path = tmp_bifur.name
            
            # Экспортируем графики во временные файлы
            exporter1 = pg.exporters.ImageExporter(self.plot_iter.plotItem)
            exporter1.parameters()['width'] = 800
            exporter1.export(iter_path)
            
            exporter2 = pg.exporters.ImageExporter(self.plot_bifur.plotItem)
            exporter2.parameters()['width'] = 800
            exporter2.export(bifur_path)
            
            # Создаем Excel файл с графиками
            wb = openpyxl.Workbook()
            
            # Лист с итерационным графиком
            ws1 = wb.active
            ws1.title = "Итерационный график"
            img1 = Image(iter_path)
            img1.width = 600
            img1.height = 400
            ws1.add_image(img1, 'A1')
            
            # Лист с бифуркационной диаграммой
            ws2 = wb.create_sheet("Бифуркация")
            img2 = Image(bifur_path)
            img2.width = 600
            img2.height = 400
            ws2.add_image(img2, 'A1')
            
            # Лист с данными (используем polars)
            ws3 = wb.create_sheet("Данные")
            
            # Добавляем данные с помощью polars
            df_data = pl.DataFrame({
                "r_parameter": [self.current_r] * 10,
                "x_sample": self.all_iterations[:10] if self.all_iterations else []
            })
            
            # Конвертируем в список для записи в Excel
            data_rows = df_data.rows()
            for i, row in enumerate(data_rows, start=1):
                for j, value in enumerate(row, start=1):
                    ws3.cell(row=i, column=j, value=value)
            
            # Сохраняем Excel файл
            wb.save(file_path)
            
            # Удаляем временные файлы
            os.unlink(iter_path)
            os.unlink(bifur_path)
            
            self.statusBar().showMessage(f"Графики и данные сохранены в Excel: {file_path}")
            
    def toggle_animation(self):
        """Запуск/остановка анимации"""
        if self.animating:
            self.timer.stop()
            self.btn_animate.setText("▶ Запустить анимацию")
            self.animating = False
        else:
            interval = int(100 / self.animation_speed.value()) # чем выше скорсть, тем меньше интервал
            self.timer.start(interval)
            self.btn_animate.setText("⏸ Пауза")
            self.animating = True
            
    def animation_step(self):
        """Шаг анимации: измерение r"""
        current_r = self.slider.value() / 100
        step = 0.01 * self.animation_direction 
        
        new_r = current_r + step
        
        # Границы изменения r
        if new_r > 4.0:
            new_r = 4.0
            self.animation_direction = -1 # меняем направление
        elif new_r < 2.5:
            new_r = 2.5
            self.animation_direction = 1
            
        self.slider.setValue(int(new_r * 100))
        

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = LogisticMapDemo()
    window.show()
    sys.exit(app.exec())